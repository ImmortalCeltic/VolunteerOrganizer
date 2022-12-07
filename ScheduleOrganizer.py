from gooey import Gooey, GooeyParser
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import os
import mkl
import json
import openpyxl as pyxl

@Gooey(program_name="Venue Grid Generator")
def parse_args():
    """ Use GooeyParser to build up the arguments we will use in our script
    Save the arguments in a default json file so that we can retrieve them
    every time we run the script.
    """
    stored_args = {}
    # get the script name without the extension & use it to build up
    # the json filename
    script_name = os.path.splitext(os.path.basename(__file__))[0]
    args_file = "{}-args.json".format(script_name)
    # Read in the prior arguments as a dictionary
    if os.path.isfile(args_file):
        with open(args_file) as data_file:
            stored_args = json.load(data_file)
    parser = GooeyParser(description='Produce a Venue Schedule')
    #Creating Dimensions for schedule grid
    #Venue Name
    parser.add_argument('Venue_Name',
                        action='store',
                        default=stored_args.get('Venue_Name'),
                        help='Specify the name of the venue')
    #Shift Time
    parser.add_argument('Shift_Time',
                        action='store',
                        default=stored_args.get('Shift_Time'),
                        help='Specify the approximate shift length')
    #Universal Shift Time Check
    parser.add_argument('--Universal_Shift_Time',
                        default=stored_args.get('Universal_Shift_Time'),
                        help='Will this venue have a single shift time?',
                        widget = 'CheckBox',
                        action='store_true')
    #Venue Positions, selected based on # of vols (if 0, then role is not needed)
    parser.add_argument('--BG_Vols',
                action='store',
                widget='IntegerField',
                default=0,
                help="Specify the number of BG volunteers that will be present")
    parser.add_argument('--FOH_Vols',
            action='store',
            widget='IntegerField',
            default=0,
            help="Specify the number of FOH volunteers that will be present")
    parser.add_argument('--GT_Vols',
            action='store',
            widget='IntegerField',
            default=0,
            help="Specify the number of Green Team volunteers that will be present")
    parser.add_argument('--Hosp_Vols',
            action='store',
            widget='IntegerField',
            default=0,
            help="Specify the number of hospitality volunteers that will be present")
    parser.add_argument('--Merch_Vols',
            action='store',
            widget='IntegerField',
            default=0,
            help="Specify the number of merchandise volunteers that will be present")
    parser.add_argument('--Stage_Vols',
            action='store',
            widget='IntegerField',
            default=0,
            help="Specify the number of staging volunteers that will be present")
    parser.add_argument('--Sec_Vols',
            action='store',
            widget='IntegerField',
            default=0,
            help="Specify the number of security volunteers that will be present")
         
    #Type of supervisors at venue
    parser.add_argument('Supervisor_Positions',
                        choices=['Beer Garden','Front of House','Green Team','Hospitality','Merchandise','Staging','Security'],
                        action='store',
                        default=stored_args.get('Supervisor_Positons'),
                        widget='Listbox',
                        nargs="+",
                        metavar="Supervisor Positons",
                        help='Specify the work positons for this venue')
    parser.add_argument('--Number_of_Shows',
                        action='store',
                        default=1,
                        widget='IntegerField',
                        help="Specify the number of shows happening at this venue")
    #Return parser function to main                    
    return parser.parse_args()

#Cell: Cell Coordinates, Sheet: Sheet Obj, Ref_Sheet: Reference Sheet Title
def insert_sheet_list(cells,sheet,ref_sheet):
        #print(str("="+ref_sheet+"!"+cells))
        sheet_call = ref_sheet+"!"
        space_formula = '&" "&'
        sheet['A1'].value = str("="+sheet_call+str(cells[0])+space_formula+sheet_call+str(cells[1])+space_formula+sheet_call+str(cells[2])+space_formula+sheet_call+str(cells[3]))
        print(sheet["A1"].value)
      


def create_cell_border(cell):
        thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
        cell.border = thin_border

def is_merged(cell,sheet):
        merged = 0
        for mergedCell in sheet.merged_cells.ranges:
                if cell.coordinate in mergedCell:
                        merged = 1
                        break
        return merged

if __name__ == '__main__':
    print("Creating Venue Schedule Grid")
    wb = pyxl.load_workbook("C:\\Users\\12508\\Documents\\ProgrammingStuff\\VolunteerOrganizer\\Test.xlsm", keep_vba=True)
    sheet = wb.active
    sheet.title = 'HelloWorld'
    wb.create_sheet("ShiftList")
    sheet_list = wb['ShiftList']
    wb.save("Test.xlsm")
    args = parse_args()
    #print(args.Venue_Name)
    #print(args.Number_of_Shows)
    #print(args.Universal_Shift_Time)
    pos_vols = [args.BG_Vols,args.FOH_Vols,args.GT_Vols,args.Hosp_Vols,args.Merch_Vols,args.Stage_Vols,args.Sec_Vols]
    pos_names = ['Beer Garden','Front of House','Green Team','Hospitality','Merchandise','Staging','Security']
    #print(pos_vols)
    #Pass all variables into excel spreadsheet generator program
    font = Font(name='Calibri',
                size=28,
                bold=True,
                italic=False,
                underline='none',
                strike=False,
                color='FF000000')
    align = Alignment(horizontal='center',
                    vertical='bottom',
                    text_rotation=0,
                    wrap_text=True,
                    shrink_to_fit=False,
                    indent=0)
    #Using Column A as base cells, for chart formattin i.e. all formatting for merging and insertion will be done in column A
    #Make Venue Name Header
    sheet['A1'] = args.Venue_Name
    sheet['A1'].font = font
    sheet['A1'].alignment = align
    #sheet['A4'] = args.Shift_Time
    #sheet['A4'].font = font
    #sheet['A4'].alignment = align
    create_cell_border(sheet['A1'])
    #create_cell_border(sheet['A4'])
    sheet.column_dimensions['A'].width = len(args.Venue_Name) #Figure out good cell width value
    sheet.row_dimensions[1].height = 30
    #sheet.row_dimensions[4].height = 30 
    #Generate columns based on number of performance days, length based on # volunteer positions and supervisor positions;
    sheet.merge_cells(start_row=1, start_column=1,end_row=1,end_column=int(args.Number_of_Shows))
    #sheet.merge_cells(start_row=4, start_column=1,end_row=4,end_column=int(args.Number_of_Shows))
    sheet['A1'].fill = PatternFill("solid", fgColor="FFC3A8")
    #For each possible volunteer position, format cells with or without a supervisor slot and with the user-inputted # of volunteer slots
    col_count = 0
    start = 4
    for i in range(int(args.Number_of_Shows)):
        counter = 0
        pos_index = 0
        #Create Date Cell
        cell = sheet.cell(row=2,column=1+col_count)
        create_cell_border(cell)
        cell_date = cell.coordinate
        wb.save("Test.xlsm")
        #Create Show Title Cell
        cell = sheet.cell(row=3,column=1+col_count)
        create_cell_border(cell)
        for vols in pos_vols:
                if int(vols) !=0:
                        #Position Name Cell
                        cell = sheet.cell(row=start+counter,column=1+col_count)
                        merged = is_merged(cell,sheet)
                        #print(merged)
                        if merged == 0:
                                #print('Inserting Position Name into row: ' + str(counter+start))
                                cell = sheet.cell(row=start+counter,column=1+col_count)
                                create_cell_border(cell)
                                cell.value = pos_names[pos_index] #Inserting Volunteer Position into cell
                                sheet.merge_cells(start_row=start+counter, start_column=1,end_row=start+counter,end_column=int(args.Number_of_Shows))
                                sheet.cell(row=start+counter,column=1+col_count).alignment = align
                                cell_position = cell.coordinate
                        else: cell_position = cell.coordinate
                        #Shift Time Cell
                        counter += 1
                        cell = sheet.cell(row=start+counter,column=1+col_count)
                        merged = is_merged(cell,sheet)
                        #Debugging Print Statement
                        #print(merged)
                        #Shift Time Cell Merge Requirement Check
                        if args.Universal_Shift_Time == True and merged == 0:
                                 #print('Merging Shift Time Cells at row: ' + counter+start)
                                create_cell_border(cell)
                                cell.value = args.Shift_Time
                                sheet.merge_cells(start_row=start+counter, start_column=1,end_row=start+counter,end_column=int(args.Number_of_Shows))
                        if args.Universal_Shift_Time == None:
                                create_cell_border(cell)
                                cell.value = args.Shift_Time    
                        sheet.cell(row=start+counter,column=1+col_count).alignment = align
                        cell_time = cell.coordinate
                        cell_venue = sheet.cell(row=1,column=1)
                        insert_sheet_list([cell_date,cell_venue.coordinate,cell_position,cell_time],sheet_list,sheet.title)
                        #Supervisor Insertion
                        if pos_names[pos_index] in args.Supervisor_Positions: 
                                #print('Inserting Supervisor into row: ' + str(start+counter+1))
                                sheet.cell(row=start+counter+1,column=1+col_count,value = pos_names[pos_index] + ' Supervisor')
                                create_cell_border(sheet.cell(row=start+counter+1,column=1+col_count))
                                counter +=1
                        #Volunteer Insertion
                        for j in range(int(vols)): 
                                sheet.cell(row=start+counter+1+j,column=1+col_count,value = 'Test')
                                #print('Inserting volunteer into row: ' +str(start+counter+1+j))
                                create_cell_border(sheet.cell(row=start+counter+1+j,column=1+col_count))
                        counter +=int(vols)+1
                pos_index +=1
        col_count +=1
    #cross-examine Supervisor_Positions in cell exclusion case

    #Create cell formatting (size, colour, etc)

    #Insert volunteer positions into designated cell

    #Insert Dates into each cell


    #sheet['A2'] = args.Number_of_Volunteers
    #sheet['A3'] = str(args.Volunteer_Supervisor_Positions)
    #sheet['A4'] = str(args.Supervisor_Positions)
    #sheet['A5'] = args.Number_of_Shows
    del wb['ShiftList1']
    wb.save("Test.xlsm")
        
    print("Done")
